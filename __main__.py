from openpyxl import load_workbook


def main() -> None:
    wb = load_workbook(
        "protected.xlsx",
        password="correct_password"
    )

    ws = wb.active
    print(ws["A1"].value)


if __name__ == "__main__":
    main()
